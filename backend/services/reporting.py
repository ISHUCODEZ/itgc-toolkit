"""
Audit deliverables — turn live CCM data into the artefacts an auditor hands over.

Two outputs are generated in-process and streamed to the browser:

  • an Excel **workpaper** — a tab per control, the full exception register with
    blank sign-off columns, run history and thresholds; the working file an
    auditor annotates and files as evidence, and
  • a PDF **ITGC continuous-monitoring report** — reliance opinion, control
    scorecard and prioritised findings; the summary that goes to management /
    the audit committee.

Both are built from the same CCM read models (reliance opinion, KPIs, trend,
exception register) so the on-screen dashboard and the filed evidence never
disagree.
"""
from __future__ import annotations
import io, time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, KeepTogether)

from . import ccm

# EY-ish palette so the deliverable looks deliberate, not default.
INK = "1A1A24"
YELLOW = "FFE600"
HEAD_FILL = "262633"
SEV_FILL = {"High": "F95D54", "Medium": "FFE600", "Low": "2DB757"}
SEV_FONT = {"High": "FFFFFF", "Medium": "1A1A24", "Low": "FFFFFF"}

TITLE = "IT General Controls — Continuous Monitoring"


# =========================================================================== #
#  Excel workpaper
# =========================================================================== #
def _hdr(ws, headers, row=1):
    fill = PatternFill("solid", fgColor=HEAD_FILL)
    font = Font(bold=True, color=YELLOW, size=10, name="Calibri")
    align = Alignment(horizontal="left", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill; cell.font = font; cell.alignment = align
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 20


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sev_cell(cell, sev):
    if sev in SEV_FILL:
        cell.fill = PatternFill("solid", fgColor=SEV_FILL[sev])
        cell.font = Font(bold=True, color=SEV_FONT[sev], size=9)
        cell.alignment = Alignment(horizontal="center")


def build_workpaper(generated_by: str = "system") -> io.BytesIO:
    opinion = ccm.reliance_opinion()
    k = ccm.kpis()
    trend = ccm.trend(limit=60)
    excs = ccm.exceptions_list(status="all")
    thresholds = ccm.thresholds_list()
    detail = ccm.current_detail()
    now = time.strftime("%Y-%m-%d %H:%M")

    wb = Workbook()
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Summary ----
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=15, color=INK)
    ws["A2"] = "Continuous Controls Monitoring — audit workpaper"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    meta = [("Report date", now), ("Prepared by", generated_by),
            ("Monitoring runs", k["total_runs"]), ("Last run", k["last_run"] or "—"),
            ("Scope", "Access (SoD), Change Management, Access Recertification")]
    r = 4
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = Font(size=10)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CONTROLS RELIANCE OPINION").font = Font(bold=True, size=11, color=INK)
    r += 1
    op_cell = ws.cell(row=r, column=1, value=f"{opinion['band_label']}  ·  score {opinion['score']}/100")
    op_cell.font = Font(bold=True, size=12, color="FFFFFF")
    op_cell.fill = PatternFill("solid", fgColor=opinion["color"].lstrip("#"))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    op_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    vc = ws.cell(row=r, column=1, value=opinion["verdict"])
    vc.font = Font(size=10, italic=True); vc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 42
    r += 1
    ws.cell(row=r, column=1, value="Drivers: " + "; ".join(
        d.replace("&gt;", ">") for d in opinion["drivers"])).font = Font(size=9, color="666666")
    r += 2

    # control scorecard
    ws.cell(row=r, column=1, value="CONTROL SCORECARD").font = Font(bold=True, size=11)
    r += 1
    _hdr(ws, ["Control", "Key metric", "Value", "Status"], row=r)
    m = k.get("latest_metrics", {})
    scard = [
        ("Segregation of Duties", "Open conflicts", m.get("sod", {}).get("conflict_count", "—"),
         "Exceptions" if m.get("sod", {}).get("conflict_count") else "Clean"),
        ("Change Management", "Compliance rate", f'{m.get("change", {}).get("compliance_rate", "—")}%',
         "Effective" if (m.get("change", {}).get("compliance_rate") or 0) >= 95 else "Exceptions"),
        ("Access Recertification", "At-risk grants", m.get("recert", {}).get("at_risk", "—"),
         "Effective" if not m.get("recert", {}).get("at_risk") else "Backlog"),
        ("Exception management", "Open / oldest (days)",
         f'{k["open_exceptions"]} / {k["oldest_open_days"]}', f'MTTR {k["mttr_days"] or "—"}d'),
    ]
    for control, metric, val, status in scard:
        r += 1
        ws.cell(row=r, column=1, value=control)
        ws.cell(row=r, column=2, value=metric)
        ws.cell(row=r, column=3, value=val)
        ws.cell(row=r, column=4, value=status)
    _autosize(ws, [30, 22, 14, 16])

    # ---- Exception Register (the heart of the workpaper) ----
    ws = wb.create_sheet("Exception Register")
    cols = ["Fingerprint", "Control", "Entity", "Rule / finding", "Severity", "Detail",
            "First seen", "Last seen", "Age (days)", "Runs seen", "Status", "Owner",
            "Remediation note", "Auditor conclusion", "Reviewed by", "Date"]
    _hdr(ws, cols)
    for i, e in enumerate(excs, start=2):
        vals = [e["fingerprint"], e["control"], e["entity"], e["rule"], e["severity"],
                e["detail"], e["first_seen_ts"], e["last_seen_ts"], e["age_days"],
                e["runs_seen"], e["status"], e.get("owner", ""), e.get("note", ""),
                "", "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(j in (4, 6, 13)))
            cell.border = border
        _sev_cell(ws.cell(row=i, column=5), e["severity"])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(2, len(excs)+1)}"
    _autosize(ws, [13, 11, 20, 26, 10, 34, 17, 17, 9, 9, 13, 14, 28, 24, 14, 12])

    # ---- per-control detail tabs ----
    _sod_tab(wb, detail["sod"], border)
    _change_tab(wb, detail["change"], border)
    _recert_tab(wb, detail["recert"], border)

    # ---- Run history ----
    ws = wb.create_sheet("Run History")
    _hdr(ws, ["Run", "Timestamp", "Triggered by", "SoD conflicts", "Change compliance %",
              "Recert at-risk", "New", "Resolved", "Persisting"])
    for i, t in enumerate(trend, start=2):
        for j, v in enumerate([t["run_id"], t["ts"], t["triggered_by"], t["sod_conflicts"],
                               t["change_compliance"], t["recert_at_risk"], t["new"],
                               t["resolved"], t["persisting"]], 1):
            ws.cell(row=i, column=j, value=v).font = Font(size=9)
    _autosize(ws, [7, 20, 13, 14, 18, 14, 8, 10, 11])

    # ---- Thresholds ----
    ws = wb.create_sheet("Alert Thresholds")
    _hdr(ws, ["ID", "Control", "Metric", "Operator", "Value", "Severity", "Enabled"])
    for i, t in enumerate(thresholds, start=2):
        for j, v in enumerate([t["id"], t["control"], t["metric"], t["operator"],
                               t["value"], t["severity"], "Yes" if t["enabled"] else "No"], 1):
            ws.cell(row=i, column=j, value=v).font = Font(size=9)
    _autosize(ws, [6, 12, 20, 10, 10, 10, 9])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _sod_tab(wb, sod_r, border):
    ws = wb.create_sheet("SoD Conflicts")
    _hdr(ws, ["User", "Rule ID", "Conflict", "Severity", "Conflicting roles", "Risk", "Control ref"])
    for i, f in enumerate(sod_r["conflicts"], start=2):
        for j, v in enumerate([f["user"], f["rule_id"], f["rule"], f["severity"],
                               " + ".join(f["conflicting_roles"]), f["risk"], f["control"]], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.alignment = Alignment(vertical="top", wrap_text=(j in (6,)))
            cell.border = border
        _sev_cell(ws.cell(row=i, column=4), f["severity"])
    _autosize(ws, [16, 9, 28, 10, 26, 50, 30])


def _change_tab(wb, change_r, border):
    ws = wb.create_sheet("Change Exceptions")
    _hdr(ws, ["Change ID", "Developer", "Deployer", "Type", "Approved", "Tested",
              "Compliant", "Exceptions"])
    for i, row in enumerate(change_r["results"], start=2):
        exc = "; ".join(f'{e["title"]} ({e["severity"]})' for e in row["exceptions"]) or "—"
        for j, v in enumerate([row["change_id"], row["developer"], row["deployer"], row["type"],
                               "Y" if row["approved"] else "N", "Y" if row["tested"] else "N",
                               "Yes" if row["compliant"] else "No", exc], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.alignment = Alignment(vertical="top", wrap_text=(j == 8))
            cell.border = border
            if j == 7 and not row["compliant"]:
                cell.fill = PatternFill("solid", fgColor="FDE7E6")
    _autosize(ws, [14, 16, 16, 11, 9, 8, 10, 44])


def _recert_tab(wb, recert_r, border):
    ws = wb.create_sheet("Recert Backlog")
    _hdr(ws, ["User", "System", "Access level", "Last reviewed", "Cadence (days)", "Status", "Days"])
    fill_map = {"Overdue": "FDE7E6", "Never reviewed": "FDE7E6", "Due soon": "FFF7D6"}
    for i, it in enumerate(recert_r["items"], start=2):
        for j, v in enumerate([it["user"], it["system"], it["access_level"], it["last_reviewed"],
                               it["cadence_days"], it["status"], it["days"]], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
            if j == 6 and it["status"] in fill_map:
                cell.fill = PatternFill("solid", fgColor=fill_map[it["status"]])
    _autosize(ws, [16, 18, 20, 14, 14, 16, 8])


# =========================================================================== #
#  PDF report
# =========================================================================== #
def _page_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#888888"))
    canvas.drawString(18 * mm, 12 * mm,
                      f"{TITLE} · generated {time.strftime('%Y-%m-%d %H:%M')}")
    canvas.drawRightString(192 * mm, 12 * mm, f"Page {doc.page}")
    canvas.setStrokeColor(colors.HexColor("#FFE600"))
    canvas.setLineWidth(2)
    canvas.line(18 * mm, 15 * mm, 192 * mm, 15 * mm)
    canvas.restoreState()


def build_pdf(generated_by: str = "system") -> io.BytesIO:
    opinion = ccm.reliance_opinion()
    k = ccm.kpis()
    trend = ccm.trend(limit=60)
    excs = ccm.exceptions_list(status="open")
    m = k.get("latest_metrics", {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=22 * mm,
                            leftMargin=18 * mm, rightMargin=18 * mm, title=TITLE)
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Title"], fontSize=18, textColor=colors.HexColor("#1A1A24"), spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9.5, textColor=colors.HexColor("#666666"), spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], fontSize=12, textColor=colors.HexColor("#1A1A24"),
                        spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("body", parent=ss["Normal"], fontSize=9.5, leading=14)
    small = ParagraphStyle("small", parent=ss["Normal"], fontSize=8, textColor=colors.HexColor("#666666"))

    el = []
    el.append(Paragraph(TITLE, h1))
    el.append(Paragraph("Continuous Controls Monitoring report &nbsp;·&nbsp; "
                        f"prepared by <b>{generated_by}</b> &nbsp;·&nbsp; {time.strftime('%d %B %Y')}", sub))

    # opinion banner
    oc = colors.HexColor(opinion["color"])
    op_tbl = Table([[Paragraph(f'<font color="white"><b>CONTROLS RELIANCE OPINION</b></font>', body)],
                    [Paragraph(f'<font color="white" size="14"><b>{opinion["band_label"]}</b>'
                               f' &nbsp; (score {opinion["score"]}/100)</font>', body)]],
                   colWidths=[174 * mm])
    op_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), oc),
                                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                                ("TOPPADDING", (0, 0), (-1, -1), 6),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    el.append(op_tbl)
    el.append(Spacer(1, 4))
    el.append(Paragraph(opinion["verdict"], body))
    el.append(Paragraph("<b>Basis for opinion:</b> " + "; ".join(
        d.replace("&gt;", "&gt;") for d in opinion["drivers"]) + ".", small))

    # scorecard
    el.append(Paragraph("Control health scorecard", h2))
    sc_data = [["Control", "Key metric", "Value", "Assessment"],
               ["Segregation of Duties", "Open conflicts",
                str(m.get("sod", {}).get("conflict_count", "—")),
                "Exceptions noted" if m.get("sod", {}).get("conflict_count") else "Clean"],
               ["Change Management", "Compliance rate",
                f'{m.get("change", {}).get("compliance_rate", "—")}%',
                "Effective" if (m.get("change", {}).get("compliance_rate") or 0) >= 95 else "Exceptions noted"],
               ["Access Recertification", "At-risk grants",
                str(m.get("recert", {}).get("at_risk", "—")),
                "Effective" if not m.get("recert", {}).get("at_risk") else "Backlog"]]
    sc = Table(sc_data, colWidths=[48 * mm, 44 * mm, 36 * mm, 46 * mm])
    sc.setStyle(_grey_table())
    el.append(sc)

    # KPI strip
    el.append(Paragraph("Monitoring metrics", h2))
    kpi_data = [["Open exceptions", "High severity", "Oldest open (days)", "MTTR (days)", "Runs"],
                [str(k["open_exceptions"]), str(k["open_by_severity"]["High"]),
                 str(k["oldest_open_days"]), str(k["mttr_days"] or "—"), str(k["total_runs"])]]
    kt = Table(kpi_data, colWidths=[35 * mm] * 5)
    kt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#262633")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FFE600")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"), ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD"))]))
    el.append(kt)

    # trend summary line
    if len(trend) >= 2:
        first, last = trend[0], trend[-1]
        el.append(Spacer(1, 6))
        el.append(Paragraph(
            f"Over {len(trend)} monitoring runs ({first['ts'][:10]} → {last['ts'][:10]}): "
            f"SoD conflicts {first['sod_conflicts']} → {last['sod_conflicts']}, "
            f"change compliance {first['change_compliance']}% → {last['change_compliance']}%, "
            f"recert backlog {first['recert_at_risk']} → {last['recert_at_risk']} at-risk grants.", small))

    # findings
    el.append(Paragraph(f"Open findings — prioritised ({len(excs)})", h2))
    if not excs:
        el.append(Paragraph("No open exceptions outstanding.", body))
    else:
        rows = [["#", "Control", "Finding", "Sev", "Age", "Owner / status"]]
        for i, e in enumerate(excs[:25], 1):
            rows.append([str(i), e["control"],
                         Paragraph(f'<b>{e["entity"]}</b> — {e["rule"]}<br/>'
                                   f'<font size="7" color="#666666">{e["detail"]}</font>', small),
                         e["severity"], f'{e["age_days"]}d',
                         Paragraph(f'{e.get("owner") or "—"}<br/>'
                                   f'<font size="7" color="#666666">{e["status"]}</font>', small)])
        ft = Table(rows, colWidths=[7 * mm, 22 * mm, 92 * mm, 14 * mm, 13 * mm, 26 * mm], repeatRows=1)
        st = [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#262633")),
              ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FFE600")),
              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
              ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
              ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]
        for i, e in enumerate(excs[:25], 1):
            c = SEV_FILL.get(e["severity"])
            if c:
                st.append(("BACKGROUND", (3, i), (3, i), colors.HexColor("#" + c)))
                st.append(("TEXTCOLOR", (3, i), (3, i),
                           colors.white if e["severity"] != "Medium" else colors.HexColor("#1A1A24")))
                st.append(("ALIGN", (3, i), (3, i), "CENTER"))
        ft.setStyle(TableStyle(st))
        el.append(ft)
        if len(excs) > 25:
            el.append(Paragraph(f"… and {len(excs) - 25} further open findings (see Excel workpaper).", small))

    el.append(Spacer(1, 10))
    el.append(Paragraph("This report is generated from continuous control tests run against the full "
                        "population. Findings are deterministic and trace to a named rule, risk and "
                        "control reference. Prepared as audit evidence; subject to reviewer sign-off.", small))

    doc.build(el, onFirstPage=_page_footer, onLaterPages=_page_footer)
    buf.seek(0)
    return buf


# =========================================================================== #
#  Shared helpers for GAM + FAIT reports
# =========================================================================== #

# GAM four GITC domains and which toolkit controls map to each
GAM_DOMAINS = [
    ("Access to programs and data", ["SoD", "Recert", "JML"],
     "Controls over who can access production systems and data, including "
     "user provisioning, access removal, and segregation of duties."),
    ("Program changes",             ["Change"],
     "Controls over the authorisation, testing, and independent deployment "
     "of changes to production systems."),
    ("Computer operations",         [],
     "Controls over batch processing, job scheduling, backup & recovery, "
     "and incident management. (Checklist-based — not population tested.)"),
    ("Program development",         [],
     "Controls over the implementation of new systems or major "
     "enhancements, including project governance and go-live approval."),
]

# SOX deficiency classification
def _classify(exc: dict) -> str:
    sev = exc.get("severity", "Low")
    age = exc.get("age_days", 0) or 0
    if sev == "High" and age >= 30:
        return "Material Weakness"
    if sev == "High":
        return "Significant Deficiency"
    if sev == "Medium":
        return "Significant Deficiency"
    return "Control Deficiency"

# 5-component FAIT finding text (auto-generated from exception data)
def _five_component(exc: dict) -> dict:
    ctrl = exc.get("control", "")
    entity = exc.get("entity", "")
    rule = exc.get("rule", "")
    detail = exc.get("detail", "")
    age = exc.get("age_days", 0) or 0

    cond_map = {
        "SoD": f"User {entity} holds toxic role combination '{rule}' in violation of the "
               f"organisation's segregation-of-duties policy.",
        "Change": f"Change {entity} failed the '{rule}' control check. {detail}",
        "Recert": f"Access grant for {entity} has not been recertified; status: {rule}.",
        "JML": f"Terminated user {entity} retains active system access: {detail}.",
    }
    criteria_map = {
        "SoD": "No single user should hold roles from both sides of a defined toxic pair "
               "(SOX 404; COBIT DSS06.03; ISO 27001 A.9.2).",
        "Change": "All production changes must be approved, tested, and deployed by a person "
                  "independent of the developer (SOX 404; COBIT BAI06).",
        "Recert": "Access grants must be formally reviewed on the defined cadence (quarterly / "
                  "semi-annual) to confirm continued appropriateness (ISO 27001 A.9.2.6).",
        "JML": "User access must be revoked within 24 hours of confirmed termination "
               "(SOX 404; ISO 27001 A.9.2.6; NIST CSF PR.AC-1).",
    }
    cause_map = {
        "SoD": "Access provisioning process did not enforce role conflict checks at the point "
               "of grant, or the SoD ruleset was not applied during periodic access reviews.",
        "Change": "Change management process did not enforce all required controls prior to "
                  "promoting the change to production.",
        "Recert": "Access review process was not completed within the required cadence, or the "
                  "grant was never subject to a formal recertification.",
        "JML": "IT provisioning team was not notified of the termination, or the off-boarding "
               "process was not followed within the required timeframe.",
    }
    effect_map = {
        "SoD": "Increases risk of unauthorised transactions, fraud, or error that goes "
               "undetected through lack of independent oversight.",
        "Change": "Increases risk of unauthorised, untested, or erroneous code reaching "
                  "production, potentially impacting financial data integrity.",
        "Recert": "Former or transferred employees may retain inappropriate access, increasing "
                  "risk of unauthorised access to financial systems.",
        "JML": f"Former employee has maintained access for {age} days post-termination, "
               f"creating unmitigated risk of unauthorised data access or manipulation.",
    }
    rec_map = {
        "SoD": "Management should revoke conflicting roles immediately, implement automated "
               "SoD enforcement at provisioning, and re-run the full population check monthly.",
        "Change": "Management should enforce the change management control checklist as a "
                  "hard gate prior to production deployment and conduct a retrospective review.",
        "Recert": "Management should complete the overdue recertification review immediately "
                  "and implement automated reminders at 30/15/0 days before cadence expiry.",
        "JML": "Management should revoke access immediately and implement an automated "
               "HR-to-IT feed to trigger access removal upon termination confirmation.",
    }
    return {
        "condition":      cond_map.get(ctrl, detail),
        "criteria":       criteria_map.get(ctrl, "Control should operate as designed."),
        "cause":          cause_map.get(ctrl, "Root cause to be determined by management."),
        "effect":         effect_map.get(ctrl, "Financial reporting risk — to be assessed."),
        "recommendation": rec_map.get(ctrl, "Management to remediate and document action plan."),
    }


# =========================================================================== #
#  GAM — GITC Summary Memo (PDF)
# =========================================================================== #
def build_gam_pdf(generated_by: str = "system") -> io.BytesIO:
    opinion = ccm.reliance_opinion()
    k       = ccm.kpis()
    trend   = ccm.trend(limit=60)
    excs    = ccm.exceptions_list(status="all")
    m       = k.get("latest_metrics", {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=22 * mm, bottomMargin=24 * mm,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            title="GITC Summary Memo")
    ss  = getSampleStyleSheet()
    c_ink    = colors.HexColor("#1A1A2E")
    c_yellow = colors.HexColor("#FFE600")
    c_muted  = colors.HexColor("#666666")
    c_head   = colors.HexColor("#1A1A2E")

    h1   = ParagraphStyle("g_h1",   parent=ss["Title"],   fontSize=16, textColor=c_ink,    spaceAfter=2)
    sub  = ParagraphStyle("g_sub",  parent=ss["Normal"],  fontSize=8.5, textColor=c_muted, spaceAfter=8)
    h2   = ParagraphStyle("g_h2",   parent=ss["Heading2"],fontSize=11, textColor=c_ink,    spaceBefore=10, spaceAfter=5)
    body = ParagraphStyle("g_body", parent=ss["Normal"],  fontSize=9.5, leading=14)
    sm   = ParagraphStyle("g_sm",   parent=ss["Normal"],  fontSize=8,  textColor=c_muted)
    bold = ParagraphStyle("g_bold", parent=ss["Normal"],  fontSize=9.5, leading=14, fontName="Helvetica-Bold")

    el = []

    # ── cover block ──────────────────────────────────────────────────────────
    cover = Table([
        [Paragraph('<font color="white"><b>EY GLOBAL AUDIT METHODOLOGY</b></font>', sm)],
        [Paragraph('<font color="white" size="15"><b>GITC Summary Memo</b></font>', h1)],
        [Paragraph(f'<font color="#AAAAAA" size="8">General IT Controls · Financial Audit IT (FAIT)</font>', sm)],
    ], colWidths=[170 * mm])
    cover.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), c_head),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    el.append(cover)
    el.append(Spacer(1, 6))

    # memo header table
    hdr_data = [
        ["Prepared by", generated_by,          "Report date", time.strftime("%d %B %Y")],
        ["Period",      "Year ended / current", "Methodology", "EY GAM — GITC testing"],
        ["Scope",       "Access · Change · Operations · Development", "Version", "1.0"],
    ]
    ht = Table(hdr_data, colWidths=[28 * mm, 62 * mm, 28 * mm, 52 * mm])
    ht.setStyle(TableStyle([
        ("FONTSIZE",  (0, 0), (-1, -1), 8.5),
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1), "Helvetica-Bold"),
        ("GRID",      (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("BACKGROUND",(0, 0), (-1, -1), colors.HexColor("#F8F8FA")),
        ("TOPPADDING",(0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    el.append(ht)

    # ── section 1: IT Risk Assessment ────────────────────────────────────────
    el.append(Paragraph("1. IT Risk Assessment", h2))
    sod_m    = m.get("sod",    {})
    chg_m    = m.get("change", {})
    rec_m    = m.get("recert", {})
    n_open_h = k["open_by_severity"].get("High", 0)
    risk_lvl = "High" if n_open_h >= 3 else ("Moderate" if n_open_h >= 1 else "Low")

    ra_data = [
        ["Risk factor", "Assessment", "Impact on GITC scope"],
        ["System complexity",        "Moderate",  "Standard GITC population testing applies"],
        ["Degree of IT reliance",    "High",      "Financial reporting relies on in-scope IT systems"],
        ["IT changes in period",     f'{chg_m.get("total","—")} changes tested', "Change control testing required"],
        ["History of deficiencies",  f'{k["open_exceptions"]} open exceptions', "Prior findings increase inherent risk"],
        ["Overall IT risk level",    risk_lvl,    "Drives depth and breadth of GITC procedures"],
    ]
    ra = Table(ra_data, colWidths=[52 * mm, 42 * mm, 76 * mm])
    ra.setStyle(_gam_table())
    el.append(ra)
    el.append(Paragraph(
        f"IT risk is assessed as <b>{risk_lvl}</b> based on {k['open_exceptions']} open "
        f"exceptions ({n_open_h} high severity) across {k['total_runs']} monitoring runs. "
        f"The testing approach is set accordingly.", body))

    # ── section 2: Relevant IT Systems ───────────────────────────────────────
    el.append(Paragraph("2. Relevant IT Systems — in scope", h2))
    sys_data = [
        ["System", "Type", "Financial process", "Hosting", "In scope", "CUECs required"],
        ["ERP (primary)", "ERP", "GL / AP / AR / Revenue", "On-premise", "Yes", "No"],
        ["HR/Payroll system", "HRIS", "Payroll", "Cloud (SaaS)", "Yes", "Yes — SOC 1 review"],
        ["Treasury system", "TMS", "Treasury / Cash", "On-premise", "Yes", "No"],
        ["IT Service Management", "ITSM", "Change management", "On-premise", "Yes", "No"],
    ]
    st = Table(sys_data, colWidths=[30 * mm, 22 * mm, 40 * mm, 26 * mm, 18 * mm, 34 * mm])
    st.setStyle(_gam_table())
    el.append(st)

    # ── section 3: GITC results by domain ────────────────────────────────────
    el.append(Paragraph("3. GITC Testing Results — by domain", h2))
    dom_data = [["GAM Domain", "Controls tested", "Exceptions", "Conclusion"]]
    for domain, controls, _ in GAM_DOMAINS:
        exc_count = sum(1 for e in excs if e["control"] in controls
                        and e["status"] not in ("resolved", "remediated"))
        tested = ", ".join(controls) if controls else "Checklist"
        conclusion = "Deficiencies noted" if exc_count else ("Not tested this run" if not controls else "Effective")
        dom_data.append([domain, tested, str(exc_count) if controls else "N/A", conclusion])
    dt = Table(dom_data, colWidths=[52 * mm, 36 * mm, 22 * mm, 60 * mm])
    ds = _gam_table()
    for i, (_, controls, _) in enumerate(GAM_DOMAINS, 1):
        exc_c = sum(1 for e in excs if e["control"] in controls
                    and e["status"] not in ("resolved","remediated"))
        if exc_c:
            ds.add("BACKGROUND", (3, i), (3, i), colors.HexColor("#FEE2E2"))
            ds.add("TEXTCOLOR",  (3, i), (3, i), colors.HexColor("#991B1B"))
        elif controls:
            ds.add("BACKGROUND", (3, i), (3, i), colors.HexColor("#F0FDF4"))
            ds.add("TEXTCOLOR",  (3, i), (3, i), colors.HexColor("#166534"))
    dt.setStyle(ds)
    el.append(dt)

    # ── section 4: exception summary ─────────────────────────────────────────
    el.append(Paragraph("4. GITC Exceptions — summary", h2))
    open_excs = [e for e in excs if e["status"] not in ("resolved", "remediated")]
    if not open_excs:
        el.append(Paragraph("No open GITC exceptions. All controls operating effectively.", body))
    else:
        ex_data = [["#", "Domain", "Entity / finding", "Sev", "Age", "Classification", "Status"]]
        for i, e in enumerate(open_excs[:20], 1):
            domain = next((d for d, c, _ in GAM_DOMAINS if e["control"] in c), e["control"])
            ex_data.append([
                str(i), domain,
                Paragraph(f'<b>{e["entity"]}</b><br/>'
                          f'<font size="7.5">{e["rule"]}</font>', sm),
                e["severity"], f'{e["age_days"]}d',
                _classify(e), e["status"]
            ])
        et = Table(ex_data, colWidths=[7*mm, 38*mm, 60*mm, 14*mm, 12*mm, 34*mm, 15*mm], repeatRows=1)
        est = _gam_table()
        for i, e in enumerate(open_excs[:20], 1):
            c = SEV_FILL.get(e["severity"])
            if c:
                est.add("BACKGROUND", (3, i), (3, i), colors.HexColor("#" + c))
                est.add("TEXTCOLOR",  (3, i), (3, i),
                        colors.white if e["severity"] != "Medium" else c_ink)
                est.add("ALIGN",      (3, i), (3, i), "CENTER")
        et.setStyle(est)
        el.append(et)

    # ── section 5: reliance conclusion ───────────────────────────────────────
    el.append(Paragraph("5. Overall GITC Reliance Conclusion", h2))
    band_map = {
        "RELIANCE":     ("Full reliance",    "#F0FDF4", "#166534"),
        "RELIANCE_EXC": ("Partial reliance", "#FEFCE8", "#854D0E"),
        "LIMITED":      ("Limited reliance", "#FFF7ED", "#9A3412"),
        "NO_RELIANCE":  ("No reliance",      "#FEF2F2", "#991B1B"),
    }
    gam_band, bg_hex, txt_hex = band_map.get(
        opinion["band"], ("Limited reliance", "#FFF7ED", "#9A3412"))

    rel_tbl = Table([[
        Paragraph(f'<font color="{txt_hex}"><b>GAM Reliance Conclusion: {gam_band}</b></font>', body),
        Paragraph(f'<font color="{txt_hex}">Score: {opinion["score"]}/100</font>', body),
    ]], colWidths=[130 * mm, 40 * mm])
    rel_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor(bg_hex)),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
    ]))
    el.append(rel_tbl)
    el.append(Spacer(1, 4))
    el.append(Paragraph(opinion["verdict"], body))
    el.append(Paragraph(
        "<b>Basis:</b> " + "; ".join(d.replace("&gt;", ">") for d in opinion["drivers"]) + ".", sm))

    # ── rollforward note ──────────────────────────────────────────────────────
    el.append(Paragraph("6. Rollforward Procedures", h2))
    el.append(Paragraph(
        "If testing was performed at an interim date, the following rollforward procedures "
        "should be completed to extend conclusions to year-end: <b>(1)</b> Inquire of management "
        "regarding significant IT changes in the rollforward period. <b>(2)</b> Inspect evidence "
        "of key controls operating in the rollforward period. <b>(3)</b> Re-run population tests "
        "if significant changes or new systems are identified. Document rollforward conclusion "
        "in this memo and update the GITC Summary in Canvas.", body))

    el.append(Spacer(1, 8))
    el.append(Paragraph(
        "This memo has been prepared in accordance with EY Global Audit Methodology (GAM) "
        "requirements for GITC testing. All exceptions are documented with condition, control "
        "domain, and deficiency classification. Subject to Engagement Leader review.", sm))

    def _gam_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        canvas.drawString(20 * mm, 12 * mm,
                          f"EY GAM — GITC Summary Memo · prepared {time.strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(190 * mm, 12 * mm, f"Page {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#FFE600"))
        canvas.setLineWidth(2)
        canvas.line(20 * mm, 15 * mm, 190 * mm, 15 * mm)
        canvas.restoreState()

    doc.build(el, onFirstPage=_gam_footer, onLaterPages=_gam_footer)
    buf.seek(0)
    return buf


def _gam_table():
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.HexColor("#FFE600")),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8.5),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8FA")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ])


# =========================================================================== #
#  GAM — GITC Workpaper (Excel)
# =========================================================================== #
def build_gam_workpaper(generated_by: str = "system") -> io.BytesIO:
    opinion  = ccm.reliance_opinion()
    k        = ccm.kpis()
    trend    = ccm.trend(limit=60)
    excs     = ccm.exceptions_list(status="all")
    detail   = ccm.current_detail()
    now      = time.strftime("%Y-%m-%d %H:%M")
    thin     = Side(style="thin", color="DDDDDD")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def gam_hdr(ws, headers, row=1):
        fill  = PatternFill("solid", fgColor="1A1A2E")
        font  = Font(bold=True, color=YELLOW, size=10, name="Calibri")
        align = Alignment(horizontal="left", vertical="center")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = fill; cell.font = font; cell.alignment = align
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
        ws.row_dimensions[row].height = 20

    # ── Tab 1: IT Risk Assessment ─────────────────────────────────────────────
    ws = wb.active; ws.title = "IT Risk Assessment"
    ws["A1"] = "EY GAM — GITC IT Risk Assessment"
    ws["A1"].font = Font(bold=True, size=14, color=INK)
    ws["A2"] = f"Prepared by: {generated_by}  |  Date: {now}  |  Methodology: EY GAM"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    gam_hdr(ws, ["Risk factor", "Assessment", "Implication", "Auditor notes"], row=4)
    m = k.get("latest_metrics", {})
    n_h = k["open_by_severity"].get("High", 0)
    risk_lvl = "High" if n_h >= 3 else ("Moderate" if n_h >= 1 else "Low")
    ra_rows = [
        ("System complexity",       "Moderate",  "Standard population testing applies", ""),
        ("Degree of IT reliance",   "High",      "Financial reporting depends on in-scope systems", ""),
        ("IT changes in period",    f'{m.get("change",{}).get("total","—")} tested', "Change domain testing required", ""),
        ("History of deficiencies", f'{k["open_exceptions"]} open',  "Increases inherent IT risk", ""),
        ("Staff turnover / JML",    "See JML tab", "Off-boarding controls tested", ""),
        ("Overall IT risk",         risk_lvl,    "Sets depth of GITC procedures", ""),
    ]
    for i, row in enumerate(ra_rows, 5):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
    _autosize(ws, [30, 20, 42, 36])

    # ── Tab 2: Relevant IT Systems ────────────────────────────────────────────
    ws = wb.create_sheet("Relevant IT Systems")
    gam_hdr(ws, ["System", "Type", "Financial process", "Hosting", "In scope",
                 "GITC domains", "CUECs", "Auditor notes"])
    sys_rows = [
        ("ERP (primary)",       "ERP",  "GL / AP / AR / Revenue", "On-premise", "Yes",
         "Access · Change · Operations", "No", ""),
        ("HR/Payroll",          "HRIS", "Payroll",                 "Cloud SaaS",  "Yes",
         "Access · JML",                "Yes — SOC 1 required", ""),
        ("Treasury",            "TMS",  "Treasury / Cash",         "On-premise", "Yes",
         "Access · Change",             "No", ""),
        ("IT Service Mgmt",     "ITSM", "Change management",       "On-premise", "Yes",
         "Change",                      "No", ""),
    ]
    for i, row in enumerate(sys_rows, 2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
    _autosize(ws, [18, 10, 26, 14, 10, 26, 22, 30])

    # ── Tab 3: GITC Scope Matrix ──────────────────────────────────────────────
    ws = wb.create_sheet("GITC Scope")
    ws["A1"] = "GITC Domain × System Scope Matrix"
    ws["A1"].font = Font(bold=True, size=13, color=INK)
    gam_hdr(ws, ["GAM Domain", "Controls tested", "Exceptions found",
                 "Conclusion", "Testing approach"], row=3)
    for i, (domain, controls, desc) in enumerate(GAM_DOMAINS, 4):
        exc_c = sum(1 for e in excs
                    if e["control"] in controls
                    and e["status"] not in ("resolved","remediated"))
        conclusion = "Deficiencies noted" if exc_c else ("Not in scope" if not controls else "Effective")
        for j, v in enumerate([domain, ", ".join(controls) or "Checklist",
                               exc_c if controls else "N/A", conclusion, desc], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
            if j == 4 and exc_c:
                cell.fill = PatternFill("solid", fgColor="FDE7E6")
            elif j == 4 and controls:
                cell.fill = PatternFill("solid", fgColor="E6F4EA")
    _autosize(ws, [38, 20, 18, 20, 54])

    # ── Tab 4: Access Testing (SoD + JML) ────────────────────────────────────
    ws = wb.create_sheet("Access — SoD")
    gam_hdr(ws, ["User", "Rule ID", "Conflict", "Severity", "Conflicting roles",
                 "Risk", "GAM domain", "Deficiency class"])
    for i, f in enumerate(detail["sod"]["conflicts"], 2):
        exc_mock = {"severity": f["severity"], "age_days": 999}
        for j, v in enumerate([f["user"], f["rule_id"], f["rule"], f["severity"],
                               " + ".join(f["conflicting_roles"]), f["risk"],
                               "Access to programs and data", _classify(exc_mock)], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
        _sev_cell(ws.cell(row=i, column=4), f["severity"])
    _autosize(ws, [16, 9, 28, 10, 28, 48, 28, 24])

    ws = wb.create_sheet("Access — JML")
    gam_hdr(ws, ["User", "Termination date", "Department", "Days since term.",
                 "Role", "System", "Severity", "GAM domain", "Deficiency class"])
    ws["A2"] = "Run JML review on the Access · JML page to populate findings here."
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    _autosize(ws, [18, 18, 16, 16, 20, 16, 10, 28, 24])

    # ── Tab 5: Change Testing ─────────────────────────────────────────────────
    _change_tab(wb, detail["change"], border)
    wb.worksheets[-1].title = "Change Testing"

    # ── Tab 6: Exception Register (GAM format) ────────────────────────────────
    ws = wb.create_sheet("Exception Register")
    cols = ["Fingerprint", "GAM Domain", "Control", "Entity", "Rule / Finding",
            "Severity", "Age (days)", "Status", "Classification",
            "Owner", "Remediation note", "Reviewed by", "Review date"]
    gam_hdr(ws, cols)
    for i, e in enumerate(excs, 2):
        domain = next((d for d, c, _ in GAM_DOMAINS if e["control"] in c), e["control"])
        vals = [e["fingerprint"], domain, e["control"], e["entity"], e["rule"],
                e["severity"], e["age_days"], e["status"], _classify(e),
                e.get("owner",""), e.get("note",""),
                e.get("reviewed_by",""), e.get("reviewed_ts","")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
        _sev_cell(ws.cell(row=i, column=6), e["severity"])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(2, len(excs)+1)}"
    _autosize(ws, [13, 28, 11, 20, 28, 10, 10, 13, 24, 14, 30, 14, 14])

    # ── Tab 7: Deficiency Evaluation ──────────────────────────────────────────
    ws = wb.create_sheet("Deficiency Evaluation")
    gam_hdr(ws, ["Finding", "GAM Domain", "Classification", "Assertion affected",
                 "Pervasive?", "Compensating control", "Management response"])
    open_excs = [e for e in excs if e["status"] not in ("resolved","remediated")]
    for i, e in enumerate(open_excs, 2):
        domain = next((d for d, c, _ in GAM_DOMAINS if e["control"] in c), e["control"])
        cls    = _classify(e)
        assert_map = {"SoD": "Occurrence / Authorisation", "Change": "Completeness / Accuracy",
                      "Recert": "Rights & obligations", "JML": "Existence / Occurrence"}
        pervasive  = "Yes" if cls == "Material Weakness" else "No"
        for j, v in enumerate([f'{e["entity"]} — {e["rule"]}', domain, cls,
                               assert_map.get(e["control"], "Multiple"), pervasive, "", ""], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
            if j == 3:
                c_map = {"Material Weakness": "F95D54", "Significant Deficiency": "FFE600",
                         "Control Deficiency": "D4EDDA"}
                cell.fill = PatternFill("solid", fgColor=c_map.get(cls, "FFFFFF"))
    _autosize(ws, [42, 28, 24, 24, 12, 28, 32])

    # ── Tab 8: GITC Summary ───────────────────────────────────────────────────
    ws = wb.create_sheet("GITC Summary")
    ws["A1"] = "GITC Summary — Overall Reliance Conclusion"
    ws["A1"].font = Font(bold=True, size=13, color=INK)
    ws["A3"] = "Overall reliance conclusion"
    ws["A3"].font = Font(bold=True, size=11)
    band_map = {"RELIANCE": "Full reliance", "RELIANCE_EXC": "Partial reliance",
                "LIMITED": "Limited reliance", "NO_RELIANCE": "No reliance"}
    ws["B3"] = band_map.get(opinion["band"], "Limited reliance")
    ws["B3"].font = Font(bold=True, size=11)
    ws["A4"] = "Score";  ws["B4"] = f'{opinion["score"]}/100'
    ws["A5"] = "Verdict"; ws["B5"] = opinion["verdict"]
    ws["B5"].alignment = Alignment(wrap_text=True)
    ws["A6"] = "Basis for conclusion"
    ws["B6"] = "; ".join(d.replace("&gt;", ">") for d in opinion["drivers"])
    ws["B6"].alignment = Alignment(wrap_text=True)
    ws["A8"] = "Prepared by"; ws["B8"] = generated_by
    ws["A9"] = "Date";        ws["B9"] = now
    ws["A10"] = "Reviewed by"; ws["B10"] = ""
    ws["A11"] = "EL sign-off"; ws["B11"] = ""
    for r in range(3, 12):
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2).font = Font(size=10)
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[5].height = 36; ws.row_dimensions[6].height = 36
    _autosize(ws, [24, 70])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# =========================================================================== #
#  FAIT — Management Letter (PDF)
# =========================================================================== #
def build_fait_pdf(generated_by: str = "system") -> io.BytesIO:
    opinion = ccm.reliance_opinion()
    k       = ccm.kpis()
    excs    = ccm.exceptions_list(status="all")
    open_e  = [e for e in excs if e["status"] not in ("resolved","remediated")]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=22 * mm, bottomMargin=24 * mm,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            title="FAIT Management Letter")
    ss   = getSampleStyleSheet()
    c_nv = colors.HexColor("#0D2B5E")   # navy for FAIT
    c_ink = colors.HexColor("#1A1A2E")
    c_y  = colors.HexColor("#FFE600")
    c_mu = colors.HexColor("#666666")
    c_MW = colors.HexColor("#FEE2E2")
    c_SD = colors.HexColor("#FEFCE8")
    c_CD = colors.HexColor("#F0FDF4")

    h1   = ParagraphStyle("f_h1",   parent=ss["Title"],   fontSize=16, textColor=c_nv, spaceAfter=4)
    sub  = ParagraphStyle("f_sub",  parent=ss["Normal"],  fontSize=9,  textColor=c_mu, spaceAfter=6)
    h2   = ParagraphStyle("f_h2",   parent=ss["Heading2"],fontSize=11, textColor=c_nv, spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("f_body", parent=ss["Normal"],  fontSize=9.5, leading=14)
    sm   = ParagraphStyle("f_sm",   parent=ss["Normal"],  fontSize=8,  textColor=c_mu)
    bold = ParagraphStyle("f_bold", parent=ss["Normal"],  fontSize=9.5, fontName="Helvetica-Bold")

    el = []

    # ── letterhead ────────────────────────────────────────────────────────────
    lh = Table([[
        Paragraph('<font color="white"><b>FINANCIAL AUDIT IT (FAIT)</b><br/>'
                  'IT General Controls — Management Letter</font>', body),
        Paragraph(f'<font color="#AAAAAA" size="8">{time.strftime("%d %B %Y")}</font>', sub),
    ]], colWidths=[130 * mm, 40 * mm])
    lh.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), c_nv),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    el.append(lh)
    el.append(Spacer(1, 8))

    # addressee
    addr = [
        ["To:",      "[Chief Financial Officer / Audit Committee]"],
        ["From:",    f"Financial Audit IT (FAIT) — {generated_by}"],
        ["Subject:", "IT General Controls — Findings and Recommendations"],
        ["Date:",    time.strftime("%d %B %Y")],
        ["Ref:",     f"ITGC-{time.strftime('%Y')}-001"],
    ]
    at = Table(addr, colWidths=[18 * mm, 152 * mm])
    at.setStyle(TableStyle([
        ("FONTSIZE",     (0, 0), (-1, -1), 9.5),
        ("FONTNAME",     (0, 0), (0, -1),  "Helvetica-Bold"),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LINEBELOW",    (0, -1), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
    ]))
    el.append(at)

    # ── executive summary ─────────────────────────────────────────────────────
    el.append(Paragraph("Executive Summary", h2))
    el.append(Paragraph(
        f"We have completed our IT General Controls (ITGC) assessment covering access management "
        f"(SoD and JML), change management, and access recertification across the in-scope IT systems. "
        f"Our testing identified <b>{len(open_e)} open exception(s)</b> — "
        f"{k['open_by_severity'].get('High',0)} high, "
        f"{k['open_by_severity'].get('Medium',0)} medium, "
        f"{k['open_by_severity'].get('Low',0)} low severity.", body))
    el.append(Spacer(1, 4))

    # opinion strip (navy for FAIT)
    op_t = Table([[
        Paragraph(f'<font color="white"><b>ITGC Reliance Opinion: '
                  f'{opinion["band_label"]}</b>  |  Score: {opinion["score"]}/100</font>', body)
    ]], colWidths=[170 * mm])
    op_t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), c_nv),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    el.append(op_t)
    el.append(Spacer(1, 4))
    el.append(Paragraph(opinion["verdict"], body))

    # deficiency summary table
    el.append(Paragraph("Deficiency Classification Summary", h2))
    mw = sum(1 for e in open_e if _classify(e) == "Material Weakness")
    sd = sum(1 for e in open_e if _classify(e) == "Significant Deficiency")
    cd = sum(1 for e in open_e if _classify(e) == "Control Deficiency")
    def_data = [
        ["Classification",         "Count", "Definition"],
        ["Material Weakness",      str(mw), "Reasonable possibility of material misstatement not being prevented or detected."],
        ["Significant Deficiency", str(sd), "Deficiency, or combination of deficiencies, less severe than material weakness but warrants attention."],
        ["Control Deficiency",     str(cd), "Design or operating effectiveness gap; not rising to significant deficiency level."],
    ]
    dft = Table(def_data, colWidths=[44 * mm, 14 * mm, 112 * mm])
    dfs = _fait_table()
    if mw: dfs.add("BACKGROUND", (0, 1), (-1, 1), c_MW)
    if sd: dfs.add("BACKGROUND", (0, 2), (-1, 2), c_SD)
    if cd: dfs.add("BACKGROUND", (0, 3), (-1, 3), c_CD)
    dft.setStyle(dfs)
    el.append(dft)

    # ── findings (5-component format) ─────────────────────────────────────────
    el.append(Paragraph(f"Detailed Findings ({len(open_e)})", h2))
    if not open_e:
        el.append(Paragraph("No open findings at the time of this report.", body))
    else:
        for i, e in enumerate(open_e[:15], 1):
            fc = _five_component(e)
            cls = _classify(e)
            bg_map = {"Material Weakness": "#FEE2E2",
                      "Significant Deficiency": "#FEFCE8",
                      "Control Deficiency": "#F0FDF4"}
            badge_bg = bg_map.get(cls, "#F8F8FA")

            finding_block = [
                [Paragraph(f'<b>Finding {i}: {e["entity"]} — {e["rule"]}</b>', bold),
                 Paragraph(f'<font size="8"><b>{cls}</b></font>', body)],
            ]
            fb = Table(finding_block, colWidths=[130 * mm, 40 * mm])
            fb.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor(badge_bg)),
                ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            el.append(fb)

            comp_data = [
                ["Condition",      fc["condition"]],
                ["Criteria",       fc["criteria"]],
                ["Cause",          fc["cause"]],
                ["Effect",         fc["effect"]],
                ["Recommendation", fc["recommendation"]],
                ["Management\nresponse", ""],
            ]
            ct = Table(comp_data, colWidths=[28 * mm, 142 * mm])
            ct.setStyle(TableStyle([
                ("FONTSIZE",     (0, 0), (-1, -1), 8.5),
                ("FONTNAME",     (0, 0), (0, -1),  "Helvetica-Bold"),
                ("VALIGN",       (0, 0), (-1, -1),  "TOP"),
                ("GRID",         (0, 0), (-1, -1),  0.4, colors.HexColor("#DDDDDD")),
                ("TOPPADDING",   (0, 0), (-1, -1),  4),
                ("BOTTOMPADDING",(0, 0), (-1, -1),  4),
                ("LEFTPADDING",  (0, 0), (-1, -1),  6),
                ("ROWBACKGROUNDS",(0, 0), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("BACKGROUND",   (0, 5), (-1, 5),   colors.HexColor("#FFFDE7")),
                ("MINROWHEIGHT", (0, 5), (-1, 5),   20),
            ]))
            el.append(ct)
            el.append(Spacer(1, 8))

        if len(open_e) > 15:
            el.append(Paragraph(f"… and {len(open_e)-15} further findings — see the FAIT Workpaper.", sm))

    # ── sign-off ──────────────────────────────────────────────────────────────
    el.append(Spacer(1, 10))
    so_data = [
        ["Prepared by", generated_by, "Date", time.strftime("%d %B %Y")],
        ["Reviewed by", "",           "Date", ""],
        ["EL sign-off", "",           "Date", ""],
    ]
    so = Table(so_data, colWidths=[28 * mm, 62 * mm, 14 * mm, 66 * mm])
    so.setStyle(TableStyle([
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("FONTNAME",     (0, 0), (0, -1),  "Helvetica-Bold"),
        ("FONTNAME",     (0, 0), (2, -1),  "Helvetica-Bold"),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
    ]))
    el.append(so)
    el.append(Spacer(1, 6))
    el.append(Paragraph(
        "This management letter is prepared by the Financial Audit IT (FAIT) team. Findings are "
        "based on ITGC testing performed against the in-scope systems and populations described above. "
        "Management is requested to provide a written response to each finding within 30 days.", sm))

    def _fait_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        canvas.drawString(20 * mm, 12 * mm,
                          f"FAIT Management Letter · prepared {time.strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(190 * mm, 12 * mm, f"Page {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#0D2B5E"))
        canvas.setLineWidth(2)
        canvas.line(20 * mm, 15 * mm, 190 * mm, 15 * mm)
        canvas.restoreState()

    doc.build(el, onFirstPage=_fait_footer, onLaterPages=_fait_footer)
    buf.seek(0)
    return buf


def _fait_table():
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#0D2B5E")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8.5),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ])


# =========================================================================== #
#  FAIT — Workpaper (Excel)
# =========================================================================== #
def build_fait_workpaper(generated_by: str = "system") -> io.BytesIO:
    opinion = ccm.reliance_opinion()
    k       = ccm.kpis()
    excs    = ccm.exceptions_list(status="all")
    detail  = ccm.current_detail()
    now     = time.strftime("%Y-%m-%d %H:%M")
    thin    = Side(style="thin", color="DDDDDD")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    NAVY = "0D2B5E"

    wb = Workbook()

    def fait_hdr(ws, headers, row=1):
        fill  = PatternFill("solid", fgColor=NAVY)
        font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        align = Alignment(horizontal="left", vertical="center")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = fill; cell.font = font; cell.alignment = align
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
        ws.row_dimensions[row].height = 20

    # ── Tab 1: Engagement Summary ─────────────────────────────────────────────
    ws = wb.active; ws.title = "Engagement Summary"
    ws["A1"] = "FINANCIAL AUDIT IT (FAIT) — ITGC Workpaper"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = f"Prepared by: {generated_by}  |  Date: {now}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    meta = [
        ("Client", "[Client name]"),
        ("Period", "[Year ended]"),
        ("Engagement", "[Engagement ref]"),
        ("Scope", "Access (SoD, JML, Recert), Change Management"),
        ("Methodology", "EY FAIT / GAM GITC testing"),
        ("Overall opinion", f'{opinion["band_label"]}  —  {opinion["score"]}/100'),
        ("Prepared by", generated_by),
        ("Reviewed by", ""),
        ("EL sign-off", ""),
    ]
    for r, (label, val) in enumerate(meta, 4):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font   = Font(size=10)
    _autosize(ws, [22, 60])

    # ── Tab 2: Control Matrix (RACM) ──────────────────────────────────────────
    ws = wb.create_sheet("Control Matrix (RACM)")
    fait_hdr(ws, ["Control ref", "Control domain", "Control objective",
                  "Risk addressed", "Financial assertion",
                  "Test approach", "Population size", "Conclusion"])
    racm = [
        ("GITC-01", "Access — SoD",
         "No user holds toxic role combinations that allow end-to-end fraudulent transactions",
         "Fraud / error through unauthorised transactions",
         "Occurrence / Authorisation",
         "Full population — automated rule engine", "All users", ""),
        ("GITC-02", "Access — JML",
         "Terminated employees have access revoked within defined timeframe",
         "Unauthorised access by former employees",
         "Existence / Rights",
         "Full population — terminated vs entitlements cross-reference", "All exits", ""),
        ("GITC-03", "Access — Recertification",
         "All access grants are formally reviewed on the defined cadence",
         "Inappropriate access persisting due to lack of review",
         "Rights & obligations",
         "Full population — aging vs cadence", "All grants", ""),
        ("GITC-04", "Change management",
         "All production changes are authorised, tested, and independently deployed",
         "Unauthorised or untested changes impacting data integrity",
         "Completeness / Accuracy",
         "Full population — change log audit", "All changes", ""),
    ]
    for i, row in enumerate(racm, 2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 40
    _autosize(ws, [12, 18, 38, 34, 22, 36, 16, 20])

    # ── Tab 3: SoD Testing ────────────────────────────────────────────────────
    ws = wb.create_sheet("SoD Testing")
    fait_hdr(ws, ["User", "Rule ID", "Conflict name", "Severity",
                  "Conflicting roles", "Risk", "Classification", "Auditor conclusion"])
    for i, f in enumerate(detail["sod"]["conflicts"], 2):
        exc_mock = {"severity": f["severity"], "age_days": 999}
        for j, v in enumerate([f["user"], f["rule_id"], f["rule"], f["severity"],
                               " + ".join(f["conflicting_roles"]), f["risk"],
                               _classify(exc_mock), ""], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
        _sev_cell(ws.cell(row=i, column=4), f["severity"])
    _autosize(ws, [16, 9, 28, 10, 28, 46, 24, 28])

    # ── Tab 4: JML Testing ────────────────────────────────────────────────────
    ws = wb.create_sheet("JML Testing")
    fait_hdr(ws, ["User", "Termination date", "Days post-term.", "Role",
                  "System", "Severity", "Classification", "Auditor conclusion"])
    ws["A2"] = "Run the JML Access Review (Access · JML page) and paste results here, or upload a terminated-user CSV."
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    _autosize(ws, [18, 18, 16, 20, 16, 10, 24, 28])

    # ── Tab 5: Change Testing ─────────────────────────────────────────────────
    _change_tab(wb, detail["change"], border)
    wb.worksheets[-1].title = "Change Testing"

    # ── Tab 6: Recert Testing ─────────────────────────────────────────────────
    _recert_tab(wb, detail["recert"], border)
    wb.worksheets[-1].title = "Recert Testing"

    # ── Tab 7: Deficiency Evaluation ──────────────────────────────────────────
    ws = wb.create_sheet("Deficiency Evaluation")
    fait_hdr(ws, ["#", "Finding", "Domain", "Classification",
                  "Condition", "Criteria", "Cause", "Effect",
                  "Recommendation", "Repeat?", "Management response"])
    open_e = [e for e in excs if e["status"] not in ("resolved","remediated")]
    for i, e in enumerate(open_e, 2):
        fc  = _five_component(e)
        cls = _classify(e)
        domain = next((d for d, c, _ in GAM_DOMAINS if e["control"] in c), e["control"])
        for j, v in enumerate([i-1, f'{e["entity"]} — {e["rule"]}', domain, cls,
                               fc["condition"], fc["criteria"], fc["cause"],
                               fc["effect"], fc["recommendation"], "No", ""], 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(size=9); cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if j == 4:
                c_map = {"Material Weakness": "F95D54",
                         "Significant Deficiency": "FFE600", "Control Deficiency": "D4EDDA"}
                cell.fill = PatternFill("solid", fgColor=c_map.get(cls, "FFFFFF"))
        ws.row_dimensions[i].height = 60
    _autosize(ws, [5, 30, 26, 24, 36, 36, 36, 36, 36, 10, 36])

    # ── Tab 8: Management Letter Draft ────────────────────────────────────────
    ws = wb.create_sheet("Management Letter Draft")
    ws["A1"] = "DRAFT — FAIT Management Letter"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A3"] = f"To: [Chief Financial Officer / Audit Committee]"
    ws["A4"] = f"From: Financial Audit IT (FAIT) — {generated_by}"
    ws["A5"] = f"Date: {time.strftime('%d %B %Y')}"
    ws["A6"] = f"Subject: IT General Controls — Findings and Recommendations"
    ws["A8"] = "EXECUTIVE SUMMARY"
    ws["A8"].font = Font(bold=True, size=11, color=NAVY)
    ws["A9"] = (f"We completed ITGC testing covering access management, change management, "
                f"and access recertification. We identified {len(open_e)} exception(s): "
                f"{sum(1 for e in open_e if _classify(e)=='Material Weakness')} material weakness, "
                f"{sum(1 for e in open_e if _classify(e)=='Significant Deficiency')} significant deficiency, "
                f"{sum(1 for e in open_e if _classify(e)=='Control Deficiency')} control deficiency.")
    ws["A9"].alignment = Alignment(wrap_text=True); ws.row_dimensions[9].height = 36
    ws["A11"] = "ITGC RELIANCE OPINION"
    ws["A11"].font = Font(bold=True, size=11, color=NAVY)
    ws["A12"] = f'{opinion["band_label"]}  (score {opinion["score"]}/100)'
    ws["A12"].font = Font(bold=True, size=12)
    ws["A13"] = opinion["verdict"]
    ws["A13"].alignment = Alignment(wrap_text=True); ws.row_dimensions[13].height = 36
    r = 15
    ws.cell(row=r, column=1, value="DETAILED FINDINGS").font = Font(bold=True, size=11, color=NAVY)
    r += 1
    for i, e in enumerate(open_e, 1):
        fc  = _five_component(e)
        cls = _classify(e)
        ws.cell(row=r, column=1, value=f"Finding {i}: {e['entity']} — {e['rule']}  [{cls}]")
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)
        r += 1
        for label, text in [("Condition", fc["condition"]), ("Criteria", fc["criteria"]),
                             ("Cause", fc["cause"]), ("Effect", fc["effect"]),
                             ("Recommendation", fc["recommendation"]),
                             ("Management response", "")]:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9)
            ws.cell(row=r, column=2, value=text).font  = Font(size=9)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 30 if label == "Management response" else 18
            r += 1
        r += 1
    _autosize(ws, [22, 100])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _grey_table():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#262633")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FFE600")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F6F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 7)])
